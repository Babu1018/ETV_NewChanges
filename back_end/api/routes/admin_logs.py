import logging
import math
import json
import io
import re
import zipfile
from datetime import datetime
from pathlib import Path
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel
from sqlalchemy.exc import OperationalError, SQLAlchemyError
from sqlalchemy.orm import Session

from app.activity_log_link import (
    mark_activity_log_deleted_for_history,
    mark_all_saved_activity_logs_deleted,
    resolve_history_status_label,
)
from app.auth.deps import get_admin_user
from app.db import get_db
from app.models.activity_log import ActivityLog
from app.models.activity_log_edit_audio import ActivityLogEditAudio
from app.models.user import User
from logger import LOG_FILE_ASR, LOG_FILE_TTS

router = APIRouter(prefix="/api/admin", tags=["admin"])
logger = logging.getLogger("ASR")

VALIDATOR_LOG_FILES = {
    "asr": LOG_FILE_ASR,
    "tts": LOG_FILE_TTS,
}

DEFAULT_PAGE_SIZE = 20
MAX_PAGE_SIZE = 100
MAX_BULK_DOWNLOAD = 50


class ActivityLogEntry(BaseModel):
    id: str
    type: str
    createdAt: str
    userEmail: str
    userName: str
    fileName: str
    downloadName: str
    language: str
    textPreview: str
    textContent: str
    originalTextContent: str = ""
    transcriptEdits: dict = {}
    hasTranscriptEdits: bool = False
    audioFormat: str
    mimeType: str
    validatorName: str | None = None
    gender: str | None = None
    speaker: str | None = None
    editRegions: list[dict] = []
    editRegionsDisplay: str = ""
    status: str = "Unsaved"
    audioUrl: str | None = None
    audioBase64: str | None = None


class ActivityLogsResponse(BaseModel):
    entries: list[ActivityLogEntry]
    page: int
    page_size: int
    total: int
    total_pages: int


class DeleteActivityLogsBody(BaseModel):
    ids: list[UUID]


class DownloadActivityLogsBody(BaseModel):
    ids: list[UUID]
    kind: str = "bundle"


def _user_display_name(user: User | None) -> str:
    if user is None:
        return "—"
    name = f"{user.firstname or ''} {user.lastname or ''}".strip()
    return name or user.email or "—"


def _iso_timestamp(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.isoformat()


def _parse_edit_regions(raw: str | None) -> list[dict]:
    if not raw:
        return []
    try:
        data = json.loads(raw)
        return data if isinstance(data, list) else []
    except json.JSONDecodeError:
        return []


def _parse_transcript_edits(raw: str | None) -> dict:
    if not raw:
        return {}
    try:
        data = json.loads(raw)
        return data if isinstance(data, dict) else {}
    except json.JSONDecodeError:
        return {}


def _edit_regions_display(regions: list[dict]) -> str:
    if not regions:
        return ""
    labels = [str(item.get("label") or "").strip() for item in regions]
    return "\n".join(label for label in labels if label)


def _row_to_entry(
    row: ActivityLog,
    user: User | None,
    db: Session,
    base_url: str | None = None,
    token: str | None = None,
    include_base64: bool = False,
) -> ActivityLogEntry:
    timestamp = row.updated_at or row.created_at
    regions = _parse_edit_regions(getattr(row, "edit_regions", None))
    edits = _parse_transcript_edits(getattr(row, "transcript_edits", None))
    email = user.email if user else (getattr(row, "user_email", None) or "—")
    name = _user_display_name(user) if user else (getattr(row, "user_name", None) or "—")
    original_text = getattr(row, "original_text_content", None) or ""

    audio_url = None
    if base_url and token:
        audio_url = f"{base_url.rstrip('/')}/api/admin/logs/{row.activity_type}/{row.id}/audio?token={token}"

    audio_base64 = None
    if include_base64 and row.audio_data:
        import base64
        audio_base64 = f"data:{row.mime_type or 'audio/wav'};base64,{base64.b64encode(row.audio_data).decode('utf-8')}"

    return ActivityLogEntry(
        id=str(row.id),
        type=row.activity_type,
        createdAt=_iso_timestamp(timestamp),
        userEmail=email,
        userName=name,
        fileName=row.file_name,
        downloadName=row.download_name,
        language=row.language,
        textPreview=row.text_preview,
        textContent=row.text_content,
        originalTextContent=original_text,
        transcriptEdits=edits,
        hasTranscriptEdits=bool(edits.get("hasEdits")),
        audioFormat=row.audio_format,
        mimeType=row.mime_type,
        validatorName=row.validator_name or "",
        gender=row.gender,
        speaker=row.speaker,
        editRegions=regions,
        editRegionsDisplay=_edit_regions_display(regions),
        status=resolve_history_status_label(row, db),
        audioUrl=audio_url,
        audioBase64=audio_base64,
    )


def _matches_search(row: ActivityLog, user: User | None, needle: str, db: Session) -> bool:
    blob = " ".join(
        [
            row.activity_type or "",
            row.file_name or "",
            row.language or "",
            row.validator_name or "",
            row.gender or "",
            row.speaker or "",
            row.text_preview or "",
            row.text_content or "",
            getattr(row, "edit_regions", "") or "",
            getattr(row, "transcript_edits", "") or "",
            getattr(row, "original_text_content", "") or "",
            getattr(row, "history_status", "") or "",
            resolve_history_status_label(row, db),
            user.email if user else "",
            _user_display_name(user),
            getattr(row, "user_email", "") or "",
            getattr(row, "user_name", "") or "",
        ]
    ).lower()
    return needle in blob


def _load_activity_entries(
    db: Session,
    activity_type: str | None,
    search: str | None,
) -> list[ActivityLogEntry]:
    needle = search.strip().lower() if search and search.strip() else None
    normalized_type = (activity_type or "all").lower()

    query = (
        db.query(ActivityLog, User)
        .outerjoin(User, User.id == ActivityLog.user_id)
        .order_by(ActivityLog.created_at.desc())
    )
    if normalized_type in ("asr", "tts"):
        query = query.filter(ActivityLog.activity_type == normalized_type)

    entries: list[ActivityLogEntry] = []
    for row, user in query.all():
        if needle and not _matches_search(row, user, needle, db):
            continue
        entries.append(_row_to_entry(row, user, db))
    return entries


def _safe_name(value: str, fallback: str = "file", max_len: int = 80) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", (value or "").strip())[:max_len]
    return cleaned or fallback


def _entry_folder_name(row: ActivityLog) -> str:
    return f"{_safe_name(row.file_name, 'log')}_{str(row.id)[:8]}"


def _text_filename(row: ActivityLog) -> str:
    return "script.txt" if row.activity_type == "tts" else "transcript.txt"


def _entry_metadata(row: ActivityLog, user: User | None, db: Session) -> dict:
    return {
        "id": str(row.id),
        "type": row.activity_type,
        "createdAt": _iso_timestamp(row.created_at),
        "updatedAt": _iso_timestamp(row.updated_at),
        "userEmail": user.email if user else (getattr(row, "user_email", None) or ""),
        "userName": _user_display_name(user) if user else (getattr(row, "user_name", None) or ""),
        "fileName": row.file_name,
        "downloadName": row.download_name,
        "language": row.language,
        "validatorName": row.validator_name or "",
        "gender": row.gender or "",
        "speaker": row.speaker or "",
        "status": resolve_history_status_label(row, db),
        "audioFormat": row.audio_format,
        "mimeType": row.mime_type,
    }


def _format_time_precise(seconds: float) -> str:
    if seconds is None or not math.isfinite(float(seconds)) or seconds < 0:
        return "0:00.00"
    value = float(seconds)
    minutes = int(value // 60)
    whole = int(value % 60)
    decimals = int(round((value - int(value)) * 100))
    if decimals >= 100:
        whole += 1
        decimals = 0
    if whole >= 60:
        minutes += whole // 60
        whole = whole % 60
    return f"{minutes}:{whole:02d}.{decimals:02d}"


def _edit_audio_filename(status: str, start_sec: float, end_sec: float, audio_ext: str) -> str:
    start = _format_time_precise(start_sec)
    end = _format_time_precise(end_sec)
    range_part = f"({start} \u2013 {end})"
    prefix_map = {
        "replaced": "edited-audio",
        "deleted": "deleted-audio",
        "selected": "added-audio",
    }
    prefix = prefix_map.get((status or "").lower(), "clip-audio")
    ext = (audio_ext or "wav").lstrip(".")
    return f"{prefix}{range_part}.{ext}"


def _result_audio_filename(audio_ext: str) -> str:
    return f"Result-audio.{(audio_ext or 'wav').lstrip('.')}"


def _load_edit_audio_rows(db: Session, row: ActivityLog) -> tuple[list, dict[str, ActivityLogEditAudio]]:
    edit_rows = (
        db.query(ActivityLogEditAudio)
        .filter(ActivityLogEditAudio.activity_log_id == row.id)
        .order_by(ActivityLogEditAudio.start_sec)
        .all()
    )
    return edit_rows, {str(item.id): item for item in edit_rows}


def _region_times_match(a_start: float, a_end: float, b_start: float, b_end: float) -> bool:
    return round(a_start, 3) == round(b_start, 3) and round(a_end, 3) == round(b_end, 3)


def _resolve_region_edit_audio(
    region: dict,
    edit_rows: list,
    edit_rows_by_id: dict[str, ActivityLogEditAudio],
) -> ActivityLogEditAudio | None:
    audio_id = region.get("correctionAudioId")
    if audio_id:
        audio_row = edit_rows_by_id.get(str(audio_id))
        if audio_row is not None:
            return audio_row

    start = float(region.get("startSec", 0))
    end = float(region.get("endSec", 0))
    for audio_row in edit_rows:
        if _region_times_match(audio_row.start_sec, audio_row.end_sec, start, end):
            return audio_row
    return None


def _edit_region_audio_rel(
    folder: str,
    region: dict,
    audio_row: ActivityLogEditAudio,
) -> str:
    clip_ext = (audio_row.audio_format or "wav").lstrip(".")
    status = str(region.get("status") or "selected")
    start = float(region.get("startSec", audio_row.start_sec))
    end = float(region.get("endSec", audio_row.end_sec))
    fname = _edit_audio_filename(status, start, end, clip_ext)
    return f"{folder}/{fname}".replace("\\", "/")


def _escape_excel_formula_str(value: str) -> str:
    return (value or "").replace('"', '""')


def _excel_hyperlink_part(path: str, display: str) -> str:
    return (
        f'HYPERLINK("{_escape_excel_formula_str(path)}",'
        f'"{_escape_excel_formula_str(display)}")'
    )


def _build_edits_column_entries(row: ActivityLog, folder: str, db: Session) -> list[dict]:
    """Each entry: {text, audio_rel?} for the Excel Edits column."""
    regions = _parse_edit_regions(getattr(row, "edit_regions", None))
    if regions:
        edit_rows, edit_rows_by_id = _load_edit_audio_rows(db, row)
        used_audio_ids: set[str] = set()
        entries: list[dict] = []
        for region in regions:
            label = str(region.get("label") or "").strip() or "—"
            audio_row = _resolve_region_edit_audio(region, edit_rows, edit_rows_by_id)
            audio_rel = None
            if audio_row is not None:
                used_audio_ids.add(str(audio_row.id))
                audio_rel = _edit_region_audio_rel(folder, region, audio_row)
            entries.append({"text": label, "audio_rel": audio_rel})
        for audio_row in edit_rows:
            if str(audio_row.id) in used_audio_ids:
                continue
            clip_ext = (audio_row.audio_format or "wav").lstrip(".")
            fname = _edit_audio_filename("clip", audio_row.start_sec, audio_row.end_sec, clip_ext)
            audio_rel = f"{folder}/{fname}".replace("\\", "/")
            entries.append({"text": fname, "audio_rel": audio_rel})
        return entries

    edits = _parse_transcript_edits(getattr(row, "transcript_edits", None))
    if edits.get("hasEdits"):
        segments = edits.get("segments") or []
        text = "".join(str(segment.get("text", "")) for segment in segments).strip()
        return [{"text": text or "Manual edits", "audio_rel": None}]
    return [{"text": "—", "audio_rel": None}]


def _format_edits_excel_cell(entries: list[dict], *, link_audio: bool) -> tuple[str, bool]:
    """Return (cell_value, is_formula)."""
    if not entries:
        return "—", False

    if not link_audio:
        lines: list[str] = []
        for entry in entries:
            text = entry["text"] or "—"
            audio_rel = entry.get("audio_rel")
            if audio_rel:
                lines.append(f"{text}\n{audio_rel}")
            else:
                lines.append(text)
        return "\n".join(lines), False

    parts: list[str] = []
    has_hyperlink = False
    for entry in entries:
        text = entry["text"] or "—"
        audio_rel = entry.get("audio_rel")
        if audio_rel:
            has_hyperlink = True
            parts.append(_excel_hyperlink_part(audio_rel, text))
        else:
            parts.append(f'"{_escape_excel_formula_str(text)}"')

    if not has_hyperlink:
        return "\n".join(entry["text"] or "—" for entry in entries), False

    if len(parts) == 1 and parts[0].startswith("HYPERLINK"):
        return f"={parts[0]}", True

    return f"={(' & CHAR(10) & ').join(parts)}", True


def _collect_entry_files(row: ActivityLog, user: User | None, db: Session) -> dict[str, bytes]:
    folder = _entry_folder_name(row)
    files: dict[str, bytes] = {}

    audio_ext = (row.audio_format or "wav").lstrip(".")
    files[f"{folder}/{_result_audio_filename(audio_ext)}"] = row.audio_data

    text_name = _text_filename(row)
    files[f"{folder}/{text_name}"] = (row.text_content or "").encode("utf-8")

    original_text = getattr(row, "original_text_content", None) or ""
    if original_text and original_text != (row.text_content or ""):
        files[f"{folder}/original.txt"] = original_text.encode("utf-8")

    regions = _parse_edit_regions(getattr(row, "edit_regions", None))
    edit_rows, edit_rows_by_id = _load_edit_audio_rows(db, row)
    used_audio_ids: set[str] = set()

    for region in regions:
        audio_row = _resolve_region_edit_audio(region, edit_rows, edit_rows_by_id)
        if audio_row is None:
            continue
        used_audio_ids.add(str(audio_row.id))
        fname = _edit_region_audio_rel(folder, region, audio_row).split("/", 1)[-1]
        files[f"{folder}/{fname}"] = audio_row.audio_data

    for audio_row in edit_rows:
        if str(audio_row.id) in used_audio_ids:
            continue
        clip_ext = (audio_row.audio_format or "wav").lstrip(".")
        fname = _edit_audio_filename("clip", audio_row.start_sec, audio_row.end_sec, clip_ext)
        files[f"{folder}/{fname}"] = audio_row.audio_data

    return files


def _order_rows(ids: list[UUID], rows: list[ActivityLog]) -> list[ActivityLog]:
    by_id = {row.id: row for row in rows}
    return [by_id[row_id] for row_id in ids if row_id in by_id]


def _edits_summary_text(row: ActivityLog) -> str:
    regions = _parse_edit_regions(getattr(row, "edit_regions", None))
    if regions:
        return _edit_regions_display(regions)
    edits = _parse_transcript_edits(getattr(row, "transcript_edits", None))
    if edits.get("hasEdits"):
        segments = edits.get("segments") or []
        text = "".join(str(segment.get("text", "")) for segment in segments).strip()
        return text or "Manual edits"
    return "—"


def _build_summary_xlsx(
    db: Session,
    rows: list[ActivityLog],
    audio_rel_paths: list[str],
    *,
    link_audio: bool = True,
    base_url: str | None = None,
    token: str | None = None,
) -> bytes:
    """Excel index matching the Users Logs UI; hyperlinks open audio after ZIP extract or stream via API."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        raise HTTPException(
            status_code=503,
            detail="Excel export needs openpyxl. Run: pip install openpyxl",
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Users Logs"

    headers = [
        "#",
        "Date & Time",
        "Type",
        "User",
        "Email",
        "File Name",
        "Language",
        "Status",
        "Script / Transcript",
        "Edits",
        "Audio (Result-audio)",
    ]
    sheet.append(headers)
    header_font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        sheet.cell(row=1, column=col_idx).font = header_font

    link_font = Font(color="0563C1", underline="single")
    wrap = Alignment(wrap_text=True, vertical="top")

    for idx, row in enumerate(rows, 1):
        user = db.query(User).filter(User.id == row.user_id).first() if row.user_id else None
        email = user.email if user else (getattr(row, "user_email", None) or "—")
        name = _user_display_name(user) if user else (getattr(row, "user_name", None) or "—")
        timestamp = _iso_timestamp(row.updated_at or row.created_at)
        activity_label = "TTS" if row.activity_type == "tts" else "ASR"
        text_content = row.text_content or ""
        folder = _entry_folder_name(row)
        edits_entries = _build_edits_column_entries(row, folder, db)
        edits_value, edits_is_formula = _format_edits_excel_cell(edits_entries, link_audio=link_audio)
        sheet.append(
            [
                idx,
                timestamp,
                activity_label,
                name,
                email,
                row.file_name,
                row.language or "—",
                resolve_history_status_label(row, db),
                text_content,
                edits_value,
                "",
            ]
        )
        row_num = idx + 1
        sheet.cell(row=row_num, column=9).alignment = wrap
        edits_cell = sheet.cell(row=row_num, column=10)
        edits_cell.alignment = wrap
        if edits_is_formula:
            edits_cell.font = link_font

        audio_rel = audio_rel_paths[idx - 1] if idx - 1 < len(audio_rel_paths) else ""
        audio_cell = sheet.cell(row=row_num, column=11)
        if base_url and token:
            audio_url = f"{base_url.rstrip('/')}/api/admin/logs/{row.activity_type}/{row.id}/audio?token={token}"
            audio_cell.value = "Open/Play audio"
            audio_cell.hyperlink = audio_url
            audio_cell.font = link_font
        elif audio_rel:
            normalized = audio_rel.replace("\\", "/")
            if link_audio:
                audio_cell.value = "Open audio"
                audio_cell.hyperlink = normalized
                audio_cell.font = link_font
            else:
                audio_cell.value = normalized
        else:
            audio_cell.value = "—"

    sheet.column_dimensions["A"].width = 5
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["C"].width = 8
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 26
    sheet.column_dimensions["F"].width = 22
    sheet.column_dimensions["G"].width = 12
    sheet.column_dimensions["H"].width = 10
    sheet.column_dimensions["I"].width = 48
    sheet.column_dimensions["J"].width = 36
    sheet.column_dimensions["K"].width = 22

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_summary_json(
    db: Session,
    rows: list[ActivityLog],
    base_url: str | None = None,
    token: str | None = None,
    include_base64: bool = False,
) -> bytes:
    entries = []
    for row in rows:
        user = db.query(User).filter(User.id == row.user_id).first() if row.user_id else None
        entry = _row_to_entry(row, user, db, base_url=base_url, token=token, include_base64=include_base64)
        entries.append(entry.dict() if hasattr(entry, "dict") else entry.model_dump())
    data = entries[0] if len(rows) == 1 else entries
    return json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")


def _zip_file_map(file_map: dict[str, bytes]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for path, data in file_map.items():
            archive.writestr(path, data)
    return buffer.getvalue()


def _build_bundle_zip(db: Session, rows: list[ActivityLog]) -> bytes:
    file_map: dict[str, bytes] = {}
    audio_rels: list[str] = []
    for row in rows:
        user = db.query(User).filter(User.id == row.user_id).first() if row.user_id else None
        audio_ext = (row.audio_format or "wav").lstrip(".")
        folder = _entry_folder_name(row)
        audio_rels.append(f"{folder}/{_result_audio_filename(audio_ext)}")
        file_map.update(_collect_entry_files(row, user, db))
    if rows:
        file_map["users_logs_summary.xlsx"] = _build_summary_xlsx(db, rows, audio_rels)
        file_map["users_logs_summary.json"] = _build_summary_json(db, rows, include_base64=False)
    return _zip_file_map(file_map)


def _build_audio_only_zip(db: Session, rows: list[ActivityLog]) -> bytes:
    file_map: dict[str, bytes] = {}
    used_names: dict[str, int] = {}
    audio_rels: list[str] = []
    for row in rows:
        base = _safe_name(row.download_name or row.file_name, "audio")
        audio_ext = (row.audio_format or "wav").lstrip(".")
        if not base.lower().endswith(f".{audio_ext}"):
            base = f"{base}.{audio_ext}"
        count = used_names.get(base, 0)
        used_names[base] = count + 1
        entry_name = base if count == 0 else f"{Path(base).stem}_{count}{Path(base).suffix}"
        file_map[entry_name] = row.audio_data
        audio_rels.append(entry_name)
    if rows:
        file_map["users_logs_summary.xlsx"] = _build_summary_xlsx(db, rows, audio_rels)
        file_map["users_logs_summary.json"] = _build_summary_json(db, rows, include_base64=False)
    return _zip_file_map(file_map)


def _load_log_row(db: Session, entry_id: UUID, activity_type: str | None = None) -> ActivityLog:
    row = db.query(ActivityLog).filter(ActivityLog.id == entry_id).first()
    if row is None:
        raise HTTPException(status_code=404, detail="Activity entry not found")
    if activity_type and row.activity_type != activity_type.lower():
        raise HTTPException(status_code=404, detail="Activity entry not found")
    return row


def _normalize_download_kind(kind: str) -> str:
    normalized = (kind or "bundle").lower()
    if normalized not in ("audio", "text", "bundle", "excel", "json"):
        raise HTTPException(status_code=400, detail="Invalid download kind")
    return normalized


@router.get("/logs", response_model=ActivityLogsResponse)
def get_application_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=MAX_PAGE_SIZE),
    search: str | None = Query(None, max_length=200),
    type: str | None = Query(None, pattern="^(all|asr|tts)$"),
    _admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    logger.debug(
        "Admin %s viewed activity logs (page=%s, page_size=%s, type=%s, search=%r)",
        _admin.email,
        page,
        page_size,
        type,
        search,
    )
    try:
        entries = _load_activity_entries(db, type, search)
    except OperationalError as exc:
        logger.exception("Failed to load activity logs (database connection)")
        raise HTTPException(status_code=503, detail="Database unavailable") from exc
    except SQLAlchemyError as exc:
        logger.exception("Failed to load activity logs")
        raise HTTPException(status_code=503, detail="Database error") from exc

    total = len(entries)
    total_pages = max(1, math.ceil(total / page_size)) if total else 1
    safe_page = min(page, total_pages)
    start = (safe_page - 1) * page_size
    page_entries = entries[start : start + page_size]

    return ActivityLogsResponse(
        entries=page_entries,
        page=safe_page,
        page_size=page_size,
        total=total,
        total_pages=total_pages,
    )


@router.get("/logs/{activity_type}/{entry_id}/audio")
def get_activity_audio(
    activity_type: str,
    entry_id: UUID,
    _admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    normalized = activity_type.lower()
    if normalized not in ("asr", "tts"):
        raise HTTPException(status_code=400, detail="Invalid activity type")

    row = db.query(ActivityLog).filter(ActivityLog.id == entry_id).first()
    if row is None or row.activity_type != normalized:
        raise HTTPException(status_code=404, detail="Activity entry not found")

    return Response(
        content=row.audio_data,
        media_type=row.mime_type,
        headers={"Content-Disposition": f'inline; filename="{row.download_name}"'},
    )


@router.get("/logs/{activity_type}/{entry_id}/edit-audio/{audio_id}")
def get_activity_edit_correction_audio(
    activity_type: str,
    entry_id: UUID,
    audio_id: UUID,
    _admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    normalized = activity_type.lower()
    if normalized not in ("asr", "tts"):
        raise HTTPException(status_code=400, detail="Invalid activity type")

    row = (
        db.query(ActivityLogEditAudio)
        .filter(
            ActivityLogEditAudio.id == audio_id,
            ActivityLogEditAudio.activity_log_id == entry_id,
        )
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Edit correction audio not found")

    log_row = db.query(ActivityLog).filter(ActivityLog.id == entry_id).first()
    if log_row is None or log_row.activity_type != normalized:
        raise HTTPException(status_code=404, detail="Activity entry not found")

    return Response(
        content=row.audio_data,
        media_type=row.mime_type,
        headers={"Content-Disposition": f'inline; filename="{row.download_name}"'},
    )


@router.delete("/logs")
def delete_activity_logs(
    body: DeleteActivityLogsBody,
    _admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    if not body.ids:
        return {"deleted": 0}

    unique_ids = list(dict.fromkeys(body.ids))
    logger.info("Admin %s deleting %s activity log(s)", _admin.email, len(unique_ids))

    try:
        deleted = (
            db.query(ActivityLog)
            .filter(ActivityLog.id.in_(unique_ids))
            .delete(synchronize_session=False)
        )
        db.commit()
    except OperationalError as exc:
        db.rollback()
        logger.exception("Failed to delete activity logs (database connection)")
        raise HTTPException(status_code=503, detail="Database unavailable") from exc
    except SQLAlchemyError as exc:
        db.rollback()
        logger.exception("Failed to delete activity logs")
        raise HTTPException(status_code=503, detail="Database error") from exc

    return {"deleted": deleted}


@router.get("/logs/{activity_type}/{entry_id}/download")
def download_activity_log_entry(
    activity_type: str,
    entry_id: UUID,
    request: Request,
    kind: str = Query("bundle", pattern="^(audio|text|bundle|excel|json)$"),
    _admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    normalized = activity_type.lower()
    if normalized not in ("asr", "tts"):
        raise HTTPException(status_code=400, detail="Invalid activity type")

    row = _load_log_row(db, entry_id, normalized)
    download_kind = _normalize_download_kind(kind)

    auth_header = request.headers.get("Authorization", "")
    token = auth_header.split(" ")[-1] if auth_header.lower().startswith("bearer ") else None

    logger.info(
        "Admin %s downloaded activity log %s (%s, kind=%s)",
        _admin.email,
        entry_id,
        normalized,
        download_kind,
    )

    if download_kind == "audio":
        return Response(
            content=row.audio_data,
            media_type=row.mime_type,
            headers={"Content-Disposition": f'attachment; filename="{row.download_name}"'},
        )

    if download_kind == "text":
        text_name = _text_filename(row)
        safe_base = _safe_name(row.file_name, "log")
        return Response(
            content=(row.text_content or "").encode("utf-8"),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{safe_base}_{text_name}"'},
        )

    if download_kind == "json":
        json_bytes = _build_summary_json(
            db,
            [row],
            base_url=str(request.base_url),
            token=token,
            include_base64=True,
        )
        folder = _entry_folder_name(row)
        return Response(
            content=json_bytes,
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{folder}.json"'},
        )

    if download_kind == "excel":
        folder = _entry_folder_name(row)
        audio_ext = (row.audio_format or "wav").lstrip(".")
        xlsx_bytes = _build_summary_xlsx(
            db,
            [row],
            [f"{folder}/{_result_audio_filename(audio_ext)}"],
            link_audio=False,
            base_url=str(request.base_url),
            token=token,
        )
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{folder}_summary.xlsx"'},
        )

    zip_bytes = _build_bundle_zip(db, [row])
    folder = _entry_folder_name(row)
    return Response(
        content=zip_bytes,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{folder}.zip"'},
    )


@router.post("/logs/download-bulk")
def download_activity_logs_bulk(
    body: DownloadActivityLogsBody,
    request: Request,
    _admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    if not body.ids:
        raise HTTPException(status_code=400, detail="No entries selected")

    unique_ids = list(dict.fromkeys(body.ids))
    if len(unique_ids) > MAX_BULK_DOWNLOAD:
        raise HTTPException(
            status_code=400,
            detail=f"Select at most {MAX_BULK_DOWNLOAD} entries to download",
        )

    download_kind = _normalize_download_kind(body.kind)
    rows = _order_rows(unique_ids, db.query(ActivityLog).filter(ActivityLog.id.in_(unique_ids)).all())
    if len(rows) != len(unique_ids):
        raise HTTPException(status_code=404, detail="One or more activity entries not found")

    auth_header = request.headers.get("Authorization", "")
    token = auth_header.split(" ")[-1] if auth_header.lower().startswith("bearer ") else None

    logger.info(
        "Admin %s bulk-downloaded %s activity log(s) (kind=%s)",
        _admin.email,
        len(rows),
        download_kind,
    )

    stamp = datetime.utcnow().strftime("%Y%m%d")
    if download_kind == "json":
        json_bytes = _build_summary_json(
            db,
            rows,
            base_url=str(request.base_url),
            token=token,
            include_base64=True,
        )
        filename = f"users_logs_{stamp}.json" if len(rows) > 1 else f"{_entry_folder_name(rows[0])}.json"
        return Response(
            content=json_bytes,
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if download_kind == "excel":
        audio_rels = []
        for row in rows:
            folder = _entry_folder_name(row)
            audio_ext = (row.audio_format or "wav").lstrip(".")
            audio_rels.append(f"{folder}/{_result_audio_filename(audio_ext)}")
        xlsx_bytes = _build_summary_xlsx(
            db,
            rows,
            audio_rels,
            link_audio=False,
            base_url=str(request.base_url),
            token=token,
        )
        filename = f"users_logs_{stamp}.xlsx" if len(rows) > 1 else f"{_entry_folder_name(rows[0])}_summary.xlsx"
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if download_kind == "audio":
        if len(rows) == 1:
            row = rows[0]
            return Response(
                content=row.audio_data,
                media_type=row.mime_type,
                headers={"Content-Disposition": f'attachment; filename="{row.download_name}"'},
            )
        zip_bytes = _build_audio_only_zip(db, rows)
        return Response(
            content=zip_bytes,
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="users_logs_audio_{stamp}.zip"'},
        )

    zip_bytes = _build_bundle_zip(db, rows)
    filename = f"users_logs_{stamp}.zip" if len(rows) > 1 else f"{_entry_folder_name(rows[0])}.zip"
    return Response(
        content=zip_bytes,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/validator-logs/{log_kind}")
def download_validator_log(
    log_kind: str,
    _admin: User = Depends(get_admin_user),
):
    normalized = log_kind.lower()
    log_path = VALIDATOR_LOG_FILES.get(normalized)
    if not log_path:
        raise HTTPException(status_code=400, detail="Invalid log type")

    path = Path(log_path)
    if not path.is_file():
        raise HTTPException(status_code=404, detail="Log file not found")

    logger.info("Admin %s downloaded validator log: %s", _admin.email, path.name)
    return FileResponse(
        path=str(path),
        media_type="text/plain; charset=utf-8",
        filename=path.name,
    )
